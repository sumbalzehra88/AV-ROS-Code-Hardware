#!/usr/bin/env python3
"""
latency_test.py

Measures and compares latency between the C++ nodes in heading_controller_pkg
and their original Python equivalents, then writes an Excel comparison table.

Two measurement types, matching what each node actually does:
  - "event"  nodes publish their output immediately inside a subscriber
             callback -> we measure ROUND-TRIP LATENCY (publish input,
             time until the matching output arrives).
  - "timer"  nodes publish on a fixed-rate timer, independent of any
             single input -> we measure output PUBLISH JITTER (how much
             the actual interval between publishes deviates from the
             nominal 1/rate period).

Does not modify any node source file -- drives everything over real
ROS 2 topics, the same way `ros2 topic pub`/`echo` would.

PREREQUISITES:
    cd /ros2_ws
    source /opt/ros/humble/setup.bash
    colcon build --packages-select heading_controller_pkg
    source install/setup.bash
    pip install openpyxl --break-system-packages   # if not already present

SETUP:
    Fill in PY_SCRIPT_PATHS below with the absolute paths to your original
    .py files. Any node left as None is skipped for the Python side (this
    is expected for heading_controller and gps_distance_velocity_node,
    which never had Python originals).

USAGE:
    # Run everything in one go (original behavior):
    python3 latency_test.py

    # Run ONE node at a time -- results accumulate into the same Excel file
    # (and a latency_results.json sidecar that's the actual source of truth):
    python3 latency_test.py --list                       # see available node names
    python3 latency_test.py --node heading_controller     # measure just this one
    python3 latency_test.py --node gps_speed_controller   # run again for another node
    ...                                                    # each run adds/updates
                                                            # that node's row in the
                                                            # same latency_comparison.xlsx

    python3 latency_test.py --output my_comparison.xlsx
"""

import json
import os

import argparse
import csv
import statistics
import subprocess
import sys
import tempfile
import time

import rclpy
from rclpy.node import Node
from std_msgs.msg import Float32, Float32MultiArray, Bool, Int32, String
from sensor_msgs.msg import NavSatFix
from geometry_msgs.msg import TwistStamped
from nav_msgs.msg import Path, Odometry
from geometry_msgs.msg import PoseStamped


# ============================================================
# SETUP -- fill in your actual Python file paths here
# ============================================================
PY_SCRIPT_PATHS = {
    "heading_controller":            "python_files/speed.py",
    "gps_distance_velocity_node":    "python_files/speed2.py",
    "gps_speed_controller":          "python_files/speed_controller.py",
    "gps_speed_controller_alert":    "python_files/speed_controller_alert.py",
    "steering_monitor":              "python_files/dynamic.py",
    "gnss_heading_node":             "python_files/ahrs_sim.py",
    "gnss_checkpoint_monitor":       "python_files/gnss_checkpoint_monitor.py",
    "steering_aligner":              "python_files/dynamic_steering_aligner.py",
    "stanley_controller":            "python_files/stanley5.py",
    "steering_aligner2":             "python_files/steering_aligner2.py",
}


# ============================================================
# Helpers
# ============================================================
def write_waypoint_csv(path, headings):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["heading"])
        for h in headings:
            w.writerow([h])


def write_checkpoint_csv(path):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["id", "name", "lat", "lon", "tolerance_m", "desired_heading_deg", "behavior"])
        w.writerow([0, "cp0", 1.0, 1.0, 2.0, "", "align_then_continue"])


def start_cpp(executable, extra_params=None):
    cmd = ["ros2", "run", "heading_controller_pkg", executable, "--ros-args"]
    if extra_params:
        for k, v in extra_params.items():
            cmd += ["-p", f"{k}:={v}"]
    return subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)


def start_py(script_path, extra_params=None):
    cmd = ["python3", script_path, "--ros-args"]
    if extra_params:
        for k, v in extra_params.items():
            cmd += ["-p", f"{k}:={v}"]
    return subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)


def stop_proc(proc):
    if proc is None:
        return
    proc.terminate()
    try:
        proc.wait(timeout=5)
    except subprocess.TimeoutExpired:
        proc.kill()
        proc.wait()


def _drain(proc):
    """Non-blocking best-effort read of whatever the process has printed so far."""
    if proc is None or proc.stdout is None:
        return "(no output captured)"
    try:
        proc.stdout.flush()
    except Exception:
        pass
    try:
        import fcntl
        import os as _os
        fd = proc.stdout.fileno()
        fl = fcntl.fcntl(fd, fcntl.F_GETFL)
        fcntl.fcntl(fd, fcntl.F_SETFL, fl | _os.O_NONBLOCK)
        return proc.stdout.read() or "(no output captured)"
    except Exception:
        return "(could not read process output)"


def _indent(text, prefix="        "):
    return "\n".join(prefix + line for line in (text or "").splitlines()) or prefix + "(empty)"


def make_path_msg():
    msg = Path()
    ps = PoseStamped()
    ps.pose.position.x = 0.0
    ps.pose.position.y = 0.0
    ps.pose.position.z = 0.0  # ref_side
    ps.pose.orientation.w = 1.0
    msg.poses.append(ps)
    return msg


def make_odom_msg():
    msg = Odometry()
    msg.pose.pose.position.x = 1.0
    msg.pose.pose.position.y = 1.0
    msg.pose.pose.orientation.w = 1.0
    msg.twist.twist.linear.x = 1.0
    return msg


# ============================================================
# Round-trip latency (event-driven nodes)
# ============================================================
def measure_round_trip(proc_start_fn, in_topic, in_type, out_topic, out_type,
                        make_msg, n_samples, extra_setup=None):
    proc = proc_start_fn()
    if proc is None:
        return None
    time.sleep(2.0)

    rclpy.init(args=None)
    node = Node("latency_probe_rt")
    pub = node.create_publisher(in_type, in_topic, 10)

    setup_pubs = []
    if extra_setup:
        for topic, mtype, factory in extra_setup:
            p = node.create_publisher(mtype, topic, 10)
            setup_pubs.append((p, factory))
            for _ in range(3):
                p.publish(factory())
                rclpy.spin_once(node, timeout_sec=0.05)

    latencies = []
    send_time = [None]

    def on_output(_msg):
        if send_time[0] is not None:
            latencies.append((time.perf_counter() - send_time[0]) * 1000.0)
            send_time[0] = None

    node.create_subscription(out_type, out_topic, on_output, 10)

    for _ in range(n_samples):
        send_time[0] = time.perf_counter()
        pub.publish(make_msg())
        deadline = time.perf_counter() + 1.0
        while time.perf_counter() < deadline and send_time[0] is not None:
            rclpy.spin_once(node, timeout_sec=0.02)
        time.sleep(0.03)

    node.destroy_node()
    rclpy.shutdown()
    output = _drain(proc)
    stop_proc(proc)

    if not latencies:
        print(f"    [debug] no round-trip samples received. Process output:\n{_indent(output)}")
        return None
    return {
        "min": min(latencies),
        "avg": statistics.mean(latencies),
        "max": max(latencies),
        "n": len(latencies),
    }


# ============================================================
# Publish-period jitter (timer-driven nodes)
# ============================================================
def measure_jitter(proc_start_fn, out_topic, out_type, duration_s, feed_topics=None):
    proc = proc_start_fn()
    if proc is None:
        return None
    time.sleep(2.0)

    rclpy.init(args=None)
    node = Node("latency_probe_jitter")
    receive_times = []
    node.create_subscription(out_type, out_topic, lambda _m: receive_times.append(time.perf_counter()), 10)

    feed_pubs = []
    if feed_topics:
        for topic, mtype, _factory in feed_topics:
            feed_pubs.append(node.create_publisher(mtype, topic, 10))

    start = time.perf_counter()
    while time.perf_counter() - start < duration_s:
        if feed_topics:
            for pub, (_, _, factory) in zip(feed_pubs, feed_topics):
                pub.publish(factory())
        rclpy.spin_once(node, timeout_sec=0.01)

    node.destroy_node()
    rclpy.shutdown()
    output = _drain(proc)
    stop_proc(proc)

    if len(receive_times) < 3:
        print(f"    [debug] only {len(receive_times)} message(s) received. Process output:\n{_indent(output)}")
        return None
    intervals_ms = [(t2 - t1) * 1000.0 for t1, t2 in zip(receive_times, receive_times[1:])]
    return {
        "avg_period": statistics.mean(intervals_ms),
        "jitter": statistics.pstdev(intervals_ms),
        "n": len(intervals_ms),
    }


# ============================================================
# Node configuration table
# ============================================================
def build_configs(waypoint_csv_path, checkpoint_csv_path):
    return [
        # --- C++-only nodes (no Python original) ---
        {
            "name": "heading_controller",
            "type": "timer",
            "cpp_exec": "heading_controller",
            "py_key": "heading_controller",
            "out_topic": "/steering_angle", "out_type": Float32,
            "duration": 3.0,
            "extra_params": {"waypoint_csv": waypoint_csv_path},
            "feed": [
                ("/jmoab/ahrs", Float32MultiArray, lambda: Float32MultiArray(data=[0.0, 0.0, 0.0])),
                ("/steering_ang", Float32, lambda: Float32(data=0.0)),
            ],
        },
        {
            "name": "gps_distance_velocity_node",
            "type": "event",
            "cpp_exec": "gps_distance_velocity_node",
            "py_key": "gps_distance_velocity_node",
            "in_topic": "/vel", "in_type": TwistStamped,
            "out_topic": "/speed_mps", "out_type": Float32,
            "make_msg": lambda: TwistStamped(),
            "samples": 20,
        },
        # --- Nodes with both C++ and Python versions ---
        {
            "name": "gps_speed_controller",
            "type": "timer",
            "cpp_exec": "gps_speed_controller",
            "py_key": "gps_speed_controller",
            "out_topic": "/pwm_a", "out_type": Int32,
            "duration": 3.0,
            "extra_params": None,
            "feed": None,
        },
        {
            "name": "gps_speed_controller_alert",
            "type": "timer",
            "cpp_exec": "gps_speed_controller_alert",
            "py_key": "gps_speed_controller_alert",
            "out_topic": "/pwm_a", "out_type": Int32,
            "duration": 3.0,
            "extra_params": None,
            "feed": None,
        },
        {
            "name": "steering_monitor",
            "type": "event",
            "cpp_exec": "steering_monitor",
            "py_key": "steering_monitor",
            "in_topic": "/steering_angle", "in_type": Float32,
            "out_topic": "/point", "out_type": Bool,
            "make_msg": lambda: Float32(data=0.0),
            "samples": 20,
        },
        {
            "name": "gnss_heading_node",
            "type": "timer",
            "cpp_exec": "gnss_heading_node",
            "py_key": "gnss_heading_node",
            "out_topic": "/jmoab/ahrs", "out_type": Float32MultiArray,
            "duration": 3.0,
            "extra_params": None,
            "feed": None,
        },
        {
            "name": "gnss_checkpoint_monitor",
            "type": "timer",
            "cpp_exec": "gnss_checkpoint_monitor",
            "py_key": "gnss_checkpoint_monitor",
            "out_topic": "/point_info", "out_type": String,
            "duration": 3.0,
            "extra_params": {"checkpoints_file": checkpoint_csv_path, "file_type": "csv"},
            "feed": [("/fix", NavSatFix, lambda: NavSatFix(latitude=1.0, longitude=1.0))],
        },
        {
            "name": "steering_aligner",
            "type": "timer",
            "cpp_exec": "steering_aligner",
            "py_key": "steering_aligner",
            "out_topic": "/steering_angle", "out_type": Float32,
            "duration": 3.0,
            "extra_params": None,
            "feed": [
                ("/aligned", Float32, lambda: Float32(data=90.0)),
                ("/jmoab_compass", Float32MultiArray, lambda: Float32MultiArray(data=[0.0, 0.0, 0.0])),
            ],
        },
        {
            "name": "stanley_controller",
            "type": "timer",
            "cpp_exec": "stanley_controller",
            "py_key": "stanley_controller",
            "out_topic": "/steering_angle", "out_type": Float32,
            "duration": 3.0,
            "extra_params": None,
            "feed": [
                ("/odometry", Odometry, make_odom_msg),
            ],
            "one_shot_setup": [("/path_points", Path, make_path_msg)],
        },
        {
            "name": "steering_aligner2",
            "type": "timer",
            "cpp_exec": "steering_aligner2",
            "py_key": "steering_aligner2",
            "out_topic": "/steering_angle", "out_type": Float32,
            "duration": 3.0,
            "extra_params": None,
            "feed": [
                ("/aligned", Float32, lambda: Float32(data=90.0)),
                ("/jmoab_compass", Float32MultiArray, lambda: Float32MultiArray(data=[0.0, 0.0, 0.0])),
            ],
        },
    ]


# ============================================================
# Run one config against both cpp and python (if available)
# ============================================================
def run_config(cfg):
    row = {"name": cfg["name"], "type": cfg["type"], "cpp": None, "py": None}

    def start_this(lang):
        if lang == "cpp":
            return lambda: start_cpp(cfg["cpp_exec"], cfg.get("extra_params"))
        py_path = PY_SCRIPT_PATHS.get(cfg["py_key"])
        if not py_path:
            return None
        return lambda: start_py(py_path, cfg.get("extra_params"))

    for lang in ("cpp", "py"):
        starter = start_this(lang)
        if starter is None:
            continue

        # For nodes needing a one-shot setup message (e.g. stanley's /path_points),
        # publish it right after the process starts, before measuring.
        one_shot = cfg.get("one_shot_setup")

        if cfg["type"] == "event":
            result = measure_round_trip(
                starter, cfg["in_topic"], cfg["in_type"], cfg["out_topic"], cfg["out_type"],
                cfg["make_msg"], cfg.get("samples", 20),
            )
            if result:
                row[lang] = result["avg"]
        else:
            if one_shot:
                # Wrap the starter to also publish the one-shot setup message.
                def wrapped_starter(starter=starter, one_shot=one_shot):
                    p = starter()
                    time.sleep(1.0)
                    rclpy.init(args=None)
                    setup_node = Node("latency_setup_probe")
                    for topic, mtype, factory in one_shot:
                        pub = setup_node.create_publisher(mtype, topic, 10)
                        for _ in range(3):
                            pub.publish(factory())
                            rclpy.spin_once(setup_node, timeout_sec=0.05)
                    setup_node.destroy_node()
                    rclpy.shutdown()
                    return p
                result = measure_jitter(wrapped_starter, cfg["out_topic"], cfg["out_type"],
                                         cfg["duration"], cfg.get("feed"))
            else:
                result = measure_jitter(starter, cfg["out_topic"], cfg["out_type"],
                                         cfg["duration"], cfg.get("feed"))
            if result:
                row[lang] = result["avg_period"]

    return row


# ============================================================
# Persistent results store (JSON sidecar) -- lets you run one node at a
# time and accumulate results across separate invocations of this script.
# ============================================================
def load_results(json_path):
    if not os.path.exists(json_path):
        return {}
    with open(json_path, "r") as f:
        return json.load(f)


def save_results(json_path, results):
    with open(json_path, "w") as f:
        json.dump(results, f, indent=2)


# ============================================================
# Excel output
# ============================================================
def write_excel(rows, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Latency Comparison"

    arial = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    headers = ["Node", "Metric Type", "Python (ms)", "C++ (ms)", "Speedup (Py / C++)", "Notes"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    metric_label = {"event": "Round-trip latency", "timer": "Publish jitter/period (avg)"}

    start_row = 2
    for i, row in enumerate(rows):
        r = start_row + i
        ws.cell(row=r, column=1, value=row["name"]).font = arial
        ws.cell(row=r, column=2, value=metric_label[row["type"]]).font = arial

        py_val = row["py"]
        cpp_val = row["cpp"]

        py_cell = ws.cell(row=r, column=3, value=py_val if py_val is not None else "N/A")
        py_cell.font = arial
        cpp_cell = ws.cell(row=r, column=4, value=cpp_val if cpp_val is not None else "N/A")
        cpp_cell.font = arial

        if py_val is not None and cpp_val is not None and cpp_val != 0:
            ws.cell(row=r, column=5, value=f"=C{r}/D{r}").font = arial
        else:
            ws.cell(row=r, column=5, value="N/A").font = arial

        notes = ""
        if py_val is None and cpp_val is not None:
            notes = "C++-only node or Python path not set"
        ws.cell(row=r, column=6, value=notes).font = arial

    total_row = start_row + len(rows)
    ws.cell(row=total_row, column=1, value="COMBINED (SUM)").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=3,
            value=f"=SUM(C{start_row}:C{total_row - 1})").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=4,
            value=f"=SUM(D{start_row}:D{total_row - 1})").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=5,
            value=f"=C{total_row}/D{total_row}").font = Font(name="Arial", bold=True)

    widths = [28, 24, 14, 14, 20, 32]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print("NOTE: This file has live formulas. Run scripts/recalc.py (from the xlsx skill) "
          "or simply open it in Excel/LibreOffice once to populate cached values.")


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="latency_comparison.xlsx")
    parser.add_argument("--results-json", default="latency_results.json",
                         help="Sidecar file that accumulates results across separate runs.")
    parser.add_argument("--node", default=None,
                         help="Measure only this one node (see --list for names). "
                              "Result is merged into the existing results file/Excel.")
    parser.add_argument("--list", action="store_true", help="List available node names and exit.")
    args = parser.parse_args()

    tmp_wp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
    tmp_wp.close()
    write_waypoint_csv(tmp_wp.name, [90])

    tmp_cp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
    tmp_cp.close()
    write_checkpoint_csv(tmp_cp.name)

    configs = build_configs(tmp_wp.name, tmp_cp.name)
    config_by_name = {c["name"]: c for c in configs}

    if args.list:
        print("Available node names:")
        for c in configs:
            print(f"  {c['name']}")
        return

    results = load_results(args.results_json)  # {name: {"type":..., "py":..., "cpp":...}}

    if args.node:
        if args.node not in config_by_name:
            print(f"Unknown node '{args.node}'. Run with --list to see valid names.", file=sys.stderr)
            sys.exit(1)
        targets = [config_by_name[args.node]]
    else:
        targets = configs

    for cfg in targets:
        print(f"Measuring: {cfg['name']} ...")
        row = run_config(cfg)
        results[cfg["name"]] = {"type": row["type"], "py": row["py"], "cpp": row["cpp"]}
        py_str = f"{row['py']:.3f}ms" if row["py"] is not None else "N/A"
        cpp_str = f"{row['cpp']:.3f}ms" if row["cpp"] is not None else "N/A"
        print(f"  -> Python: {py_str}   C++: {cpp_str}")

    save_results(args.results_json, results)

    # Rebuild the full row list in canonical config order, including only
    # nodes that have actually been measured at least once.
    rows = []
    for c in configs:
        if c["name"] in results:
            r = results[c["name"]]
            rows.append({"name": c["name"], "type": r["type"], "py": r["py"], "cpp": r["cpp"]})

    write_excel(rows, args.output)
    print(f"Results file: {args.results_json} ({len(rows)} node(s) recorded so far)")


if __name__ == "__main__":
    main()